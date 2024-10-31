from bs4 import BeautifulSoup  # 网页分析， 爬取数据
import re  # 正则表达式， 文字匹配
import urllib.request, urllib.error  # 指定URL，获取网页数据
from openpyxl import Workbook  # 进行excel操作
import sqlite3  # 进行SQLite数据操作

# 创建正则规则对象
findLink = re.compile(r'<a href="(.*?)">')  # 影片链接规则
findImg = re.compile(r'<img.*src="(.*?)"', re.S)  # re.S: 让换行符包含在字符中 图片链接
findTitle = re.compile(r'<span class="title">(.*)</span>')  # 影片名
findRating = re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')  # 评分
findJude = re.compile(r'<span>(\d*)人评价</span>')  # 评分人数
findInq = re.compile(r'<span class="inq">(.*)</span>')  # 一句话评价
findP = re.compile(r'<p class="">(.*?)</p>', re.S)  # 影片相关内容


def askURL(url: str) -> str:
    """@爬取一页数据"""
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    }  # 为装自己为浏览器，绕开反爬虫程序
    req = urllib.request.Request(url, headers=header)
    response = urllib.request.urlopen(req)
    html = ''
    try:
        html = response.read().decode('utf-8')
        # print(html)
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html


def getData(baseurl) -> list:
    """爬取网页数据"""
    data_list = []
    for i in range(0, 10):
        url = baseurl + str(i * 25)
        htl = askURL(url)  # 保存获取到的网页源码
        soup = BeautifulSoup(htl, "html.parser")
        for item in soup.find_all('div', class_='item'):  # 查找符合要求的字符串，形成列表
            d = []  # 保存一部电影的所有信息
            item = str(item)
            # print(item)
            # break
            # 获取影片详情链接
            link = re.findall(findLink, item)[0]  # 查早指定字符串
            d.append(link)
            # print(link)
            # 获取影片图片链接
            img = re.findall(findImg, item)[0]
            d.append(img)
            # 获取影片名
            title = re.findall(findTitle, item)
            if len(title) > 1:
                zh_title = title[0]  # 中文名
                d.append(zh_title)
                fren_title = title[1].replace('/', '')  # 外文名
                d.append(fren_title.strip())
            else:
                d.extend([title[0], ' '])  # 外文名留空
            # 获取影片评分
            rate = re.findall(findRating, item)[0]
            d.append(float(rate))
            # 获取影片评价人数
            jude = re.findall(findJude, item)[0]
            d.append((int(jude)))
            # 获取影片概述
            inq = re.findall(findInq, item)
            if len(inq) != 0:
                inq = inq[0].replace('。', '')
                d.append(inq)
            else:
                d.append(' ')
            # 获取影片详情
            p = re.findall(findP, item)[0]
            p = re.sub('<br(\s+)?/>(\s+)?', '', p)
            p = re.sub('\s+', ' ', p)
            d.append(p.strip())
            # 将处理好的一部电影数据放入data_list
            data_list.append(d)

    return data_list


def saveData(data_list: list, savepath:str):
    """@保存数据"""
    wb = Workbook()  # 创建workbook对象
    sheet = wb.active  # 创建工作表
    col = ('电影详情链接', '图片链接', '影片中文名', '影片外文名', '评分', '评分人数', '概况', '相关信息')
    for i in range(0, 8):
        sheet.cell(1, i+1, col[i])
    for i in range(0, 250):
        # print(f'第{i+1}条')
        da = data_list[i]
        for j in range(0, 8):
            sheet.cell(i+2, j+1, da[j])

    wb.save(savepath)


if __name__ == '__main__':
    Url = "https://movie.douban.com/top250?start="
    data = getData(Url)
    print(*data, sep='\n')
    saveData(data, '豆瓣电影Top250.xlsx')
    print('爬取完成')
