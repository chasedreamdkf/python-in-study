import re
import tkinter as tk
import urllib.error
import urllib.request

from bs4 import BeautifulSoup
from openpyxl import Workbook

import datavisualization  # 将朱正勇的可视化代码修改为一个模块导入主程序，在主程序中输出可视化结果

# 创建正则规则对象
findLink = re.compile(r'href="(.*?)"')  # 课程视频链接
findImg = re.compile(r'src="(.*?)"')  # 课程封面图片链接
findTitle = re.compile(r'title="(.*?)"')  # 课程名称
findRate = re.compile(r'好评率(\d*.?)')  # 课程好评率
findNum = re.compile(r'(\d*.?)人报名')  # 报名人数


def askURL(url: str) -> str:
    """@获取网页源码"""
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    }  # 伪装自己为浏览器，绕开反爬虫程序
    req = urllib.request.Request(url, headers=header)
    html = ""
    try:
        response = urllib.request.urlopen(req)
        html = response.read().decode('utf-8')
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    else:
        print(html)
    return html


def getDate(h: str) -> list:
    """@获取课程数据"""
    data = []
    soup = BeautifulSoup(h, "html.parser")
    for item in soup.find_all("section", class_="course-card-expo-wrapper"):
        # print(item.prettify())
        item = str(item)
        link = "https://ke.qq.com/" + re.findall(findLink, item)[0]
        img = re.findall(findImg, item)[0]
        title = re.findall(findTitle, item)[0]
        rate = re.findall(findRate, item)
        if len(rate):  # 部分课程未进行学员打分统计
            rate = rate[0]
        else:
            rate = '0%'
        num = re.findall(findNum, item)[0]
        if num[-1] == '万':
            num = int(num[:-1]) * 10000
        else:
            num = int(num)
        # print(num)
        data.append([title, link, img, rate, num])
    return data


def savaData(data: list, savepath: str):
    """@将数据保存到excel表格中"""
    wb = Workbook()  # 创建workbook工作对象
    sheet = wb.active  # 创建工作表
    col = ('课程名称', '课程链接', '课程封面链接', '课程好评率', '课程报名人数')
    for i in range(0, 5):
        sheet.cell(1, i + 1, col[i])
    for i in range(0, len(data)):
        # print(f'第{i+1}条')
        d = data[i]
        for j in range(0, 5):
            sheet.cell(i + 2, j + 1, d[j])

    wb.save(savepath)


def course_recommend(data: list, root: tk.Tk):
    """@ 课程推荐，在新窗口展示"""
    data.sort(key=lambda x: x[-1], reverse=True)
    d = data[:50]
    d.sort(key=lambda x: int(x[-2][0:-1]), reverse=True)
    d1 = d[:10]
    # print(*d1, sep='\n')
    new_face = tk.Toplevel(root)
    new_face.geometry("700x850")
    new_face.title("课程推荐")
    txt = tk.Text(new_face)
    txt.place(rely=0, relheight=1.0, relwidth=1.0)
    i = 1
    for line in d1:
        course_name = line[0]
        course_link = line[1]
        course_num = line[-1]
        s = ("第%d个课程\n课 程 名 称 :\t%s\n课 程 链 接 :\t%s\n课程报名人数:\t%d\n\n\n" %
             (i, course_name, course_link, course_num))
        txt.insert(tk.END, s)
        i += 1
    txt.insert(tk.END, "基于报名人数和好评率综合推荐")
    close_but = tk.Button(new_face, text="关闭", command=new_face.destroy, height=2, width=12)
    close_but.place(relx=0.85, rely=0.01)


def course_seek(root: tk.Toplevel, inp: tk.Entry):
    """@ 课程查找"""
    key_word = inp.get() or "未输入"
    txt = tk.Text(root)
    txt.place(relx=0.1, rely=0.2)
    if key_word == "未输入":
        txt.insert(tk.END, "请输入关键字以查询课程")
    else:
        flag = 1
        for d in da:
            if key_word in d[0]:
                course_name = d[0]
                course_link = d[1]
                course_num = d[-1]
                s = ("课 程 名 称 :\t%s\n课 程 链 接 :\t%s\n课程报名人数:\t%d\n\n\n" %
                     (course_name, course_link, course_num))
                txt.insert(tk.END, s)
                flag = 0
        if flag:
            txt.insert(tk.END, "抱歉! 未查到您要找的课程, 请重新输入关键字查询!")


def seek_face(root: tk.Tk):
    """@ 课程搜索界面"""
    new_face = tk.Toplevel(root)
    new_face.geometry("700x800")
    new_face.title("课程搜索")
    label = tk.Label(new_face, text="课程搜索", font=("华文新魏", 30))
    label.place(relx=0.35, rely=0.01)
    label1 = tk.Label(new_face, text="请输入关键词搜索课程", font=("SimHei", 12))
    label1.place(relx=0.25, rely=0.12)
    input1 = tk.Entry(new_face, textvariable=tk.StringVar())
    input1.place(relx=0.25, rely=0.15)
    seek_but = tk.Button(new_face, text="搜索课程", height=1, width=10, command=lambda: course_seek(new_face, input1))
    seek_but.place(relx=0.5, rely=0.14)
    close_but = tk.Button(new_face, text="关闭", command=new_face.destroy, height=2, width=12)
    close_but.place(relx=0.85, rely=0.9)


def interface(text: str):
    """@用户主界面"""
    root = tk.Tk()  # 创建界面对象
    root.geometry("700x400")  # 设置界面大小
    label = tk.Label(root, text=text, height=1, width=30, font=("华文新魏", 30))
    label.place(relx=0.2, height=80, width=430)
    txt = tk.Text(root)
    txt.place(rely=0.95, relx=0.25, relheight=0.1, relwidth=0.5)
    txt.insert(tk.END, "本程序获取课程来自腾讯课堂(https://ke.qq.com/)")
    bt1 = tk.Button(root, text="推荐课程", height=2, width=15, command=lambda: course_recommend(da, root))
    bt1.place(relx=0.33, rely=0.2)
    visual_but = tk.Button(root, text="数据饼图", height=2, width=15, command=datavisualization.show)
    visual_but.place(relx=0.53, rely=0.2)
    seek_but = tk.Button(root, text="搜索课程", height=2, width=15, command=lambda: seek_face(root))
    seek_but.place(relx=0.33, rely=0.35)
    quit_but = tk.Button(root, text="退出程序", height=2, width=15, command=root.destroy)
    quit_but.place(relx=0.53, rely=0.35)
    root.mainloop()


if __name__ == '__main__':
    # Url = "https://www.icourse163.org/channel/2001.htm"
    # html = askURL(Url)
    """ 因为js异步渲染的关系, 无法爬取到完整源码, 
        所以没有使用askURL函数来获取， 只能直接复制网页源码存放到txkt.html文件中"""
    with open('./txkt.html', 'r', encoding='utf-8') as f:
        html = f.read()
    da = getDate(html)
    # print(*da, sep='\n')
    savaData(da, "./腾讯课堂课程数据.xlsx")
    da.sort(key=lambda x: x[-1], reverse=True)
    print('爬取成功!\n启动用户界面', end='')
    interface(text="在线课程评价与推荐系统")
